"""CSV/JSON/XLSX exporters: public column shape, formatting, idempotency."""
from __future__ import annotations

import csv
import json
from datetime import date, datetime

from config.schemas import ArticleRaw, Company, FundingRound, FundingStage, Investor
from pipeline.exporter import (
    PUBLIC_COLUMNS,
    _format_amount,
    _format_stage,
    _normalize_usd,
    export_csv,
    export_json,
    export_xlsx,
)
from pipeline.storage import Storage


def _seed(tmp_path) -> Storage:
    s = Storage(db_path=tmp_path / "ex.db")
    s.upsert_article(ArticleRaw(
        source="inc42", url="https://inc42.com/a", title="t", text="x",
        fetched_at=datetime(2026, 4, 17),
    ))
    s.upsert_article(ArticleRaw(
        source="entrackr", url="https://entrackr.com/a", title="t", text="x",
        fetched_at=datetime(2026, 4, 17),
    ))
    s.upsert_round(FundingRound(
        round_id="r1" * 8,
        company=Company(name="Hocco", sector="Food"),
        stage=FundingStage.SERIES_C,
        amount_usd=12_000_000,
        announced_on=date(2026, 4, 17),
        investors=[Investor(name="Sixth Sense", lead=True), Investor(name="Accel")],
        sources=["https://inc42.com/a"],
        confidence=0.95,
        summary="Hocco raised Series C.",
    ), source_urls=["https://inc42.com/a", "https://entrackr.com/a"], dedup=False)
    s.upsert_round(FundingRound(
        round_id="r2" * 8,
        company=Company(name="TraqCheck"),
        stage=FundingStage.SEED,
        amount_usd=500_000,
        announced_on=date(2026, 4, 14),
        sources=["https://inc42.com/a"],
        confidence=0.40,
    ), source_urls=["https://inc42.com/a"], dedup=False)
    return s


def test_csv_has_public_columns_only(tmp_path) -> None:
    s = _seed(tmp_path)
    out = tmp_path / "rounds.csv"
    n = export_csv(s, out)
    assert n == 2
    with out.open("r", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    # Exact column list, no round_id / extraction_method / created_at / updated_at
    assert list(rows[0].keys()) == PUBLIC_COLUMNS
    for bad in ("round_id", "extraction_method", "created_at", "updated_at"):
        assert bad not in rows[0]
    # Ordered by confidence desc
    assert rows[0]["company_name"] == "Hocco"
    assert rows[1]["company_name"] == "TraqCheck"
    assert set(rows[0]["investors"].split(" | ")) == {"Sixth Sense", "Accel"}
    assert rows[0]["lead_investor"] == "Sixth Sense"
    assert rows[0]["sector"] == "Food"
    # Stage is pretty-formatted
    assert rows[0]["stage"] == "Series C"
    # Amount formatted as string
    assert "$12M" in rows[0]["amount"]


def test_json_has_arrays_and_raw_fields(tmp_path) -> None:
    s = _seed(tmp_path)
    out = tmp_path / "rounds.json"
    n = export_json(s, out)
    assert n == 2
    data = json.loads(out.read_text(encoding="utf-8"))
    assert isinstance(data[0]["investors"], list)
    assert set(data[0]["investors"]) == {"Sixth Sense", "Accel"}
    assert isinstance(data[0]["sources"], list)
    assert len(data[0]["sources"]) == 2
    # Dashboard needs raw stage + numeric amount_usd
    assert data[0]["stage_raw"] == "series_c"
    assert data[0]["amount_usd"] == 12_000_000
    assert data[0]["stage"] == "Series C"


def test_min_confidence_filter(tmp_path) -> None:
    s = _seed(tmp_path)
    out = tmp_path / "hi.csv"
    n = export_csv(s, out, min_confidence=0.8)
    assert n == 1


def test_export_is_idempotent(tmp_path) -> None:
    s = _seed(tmp_path)
    out = tmp_path / "x.csv"
    assert export_csv(s, out) == 2
    first = out.read_text(encoding="utf-8")
    assert export_csv(s, out) == 2
    second = out.read_text(encoding="utf-8")
    assert first == second


def test_xlsx_writes_workbook(tmp_path) -> None:
    from openpyxl import load_workbook

    s = _seed(tmp_path)
    out = tmp_path / "rounds.xlsx"
    n = export_xlsx(s, out)
    assert n == 2
    wb = load_workbook(out)
    ws = wb["rounds"]
    header = [c.value for c in ws[1]]
    assert header == PUBLIC_COLUMNS
    assert ws.max_row == 3
    conf_col = PUBLIC_COLUMNS.index("confidence") + 1
    assert isinstance(ws.cell(row=2, column=conf_col).value, float)


def test_format_stage() -> None:
    assert _format_stage("series_a") == "Series A"
    assert _format_stage("series_e_plus") == "Series E+"
    assert _format_stage("pre_seed") == "Pre-Seed"
    assert _format_stage(None) == ""
    assert _format_stage("") == ""


def test_normalize_usd_handles_millions_shorthand() -> None:
    # LLM emits 8 meaning $8M
    assert _normalize_usd(8.0) == 8_000_000
    # LLM emits 18072289 meaning $18M absolute
    assert _normalize_usd(18_072_289) == 18_072_289
    assert _normalize_usd(None) is None
    assert _normalize_usd(0) is None


def test_format_amount_usd() -> None:
    assert _format_amount(8.0, 8.0, "USD") == "$8M"
    assert _format_amount(1_200_000_000, None, "USD") == "$1.2B"
    assert _format_amount(500_000, None, "USD") == "$500K"


def test_format_amount_inr_shows_both() -> None:
    out = _format_amount(18_072_289, 1_500_000_000, "INR")
    assert "₹150 Cr" in out
    assert "$18" in out


def test_format_amount_inr_sub_crore() -> None:
    # 50 lakh = 5M INR
    out = _format_amount(60_000, 5_000_000, "INR")
    assert "₹50 L" in out
    # 1 crore = 10M INR
    out2 = _format_amount(120_000, 10_000_000, "INR")
    assert "Cr" in out2


def test_format_amount_empty_when_missing() -> None:
    assert _format_amount(None, None, None) == ""
