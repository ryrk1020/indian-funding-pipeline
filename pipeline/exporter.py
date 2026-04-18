"""Flat, human-readable exports of the funding_rounds table.

CSV + JSON + XLSX are idempotent: each run writes a fresh snapshot. Sheets is
handled in `pipeline.sheets_export`.

Public row shape (what reaches CSV / JSON / XLSX / Sheets / dashboard):
  company_name, sector, stage, amount, amount_usd, announced_on,
  lead_investor, investors, sources, confidence, summary

Internal fields (round_id, created_at, updated_at, extraction_method) stay in
the DB but are deliberately stripped before export — they're implementation
detail, not analytic data.

`build_public_rows` additionally returns `stage_raw` and `amount_usd_raw` on
each row for the dashboard (so filters can work with raw enums / numbers
while the table shows formatted strings).
"""
from __future__ import annotations

import csv
import json
from datetime import date, datetime
from pathlib import Path
from typing import Any

from pipeline.storage import Storage

# Columns that go out publicly (CSV / XLSX / Sheets). Dropped vs the DB:
#   round_id, extraction_method, created_at, updated_at — internal only.
PUBLIC_COLUMNS: list[str] = [
    "company_name",
    "sector",
    "stage",
    "amount",
    "amount_usd",
    "announced_on",
    "lead_investor",
    "investors",
    "sources",
    "confidence",
    "summary",
]

# Kept as a back-compat alias for any external imports.
EXPORT_COLUMNS = PUBLIC_COLUMNS

_STAGE_LABELS: dict[str, str] = {
    "pre_seed": "Pre-Seed",
    "seed": "Seed",
    "pre_series_a": "Pre-Series A",
    "series_a": "Series A",
    "series_b": "Series B",
    "series_c": "Series C",
    "series_d": "Series D",
    "series_e_plus": "Series E+",
    "bridge": "Bridge",
    "debt": "Debt",
    "grant": "Grant",
    "ipo": "IPO",
    "acquisition": "Acquisition",
    "undisclosed": "Undisclosed",
}


def _format_stage(raw: str | None) -> str:
    if not raw:
        return ""
    return _STAGE_LABELS.get(raw, raw.replace("_", " ").title())


def _normalize_usd(value: float | None) -> float | None:
    """Coerce `amount_usd` to absolute dollars.

    The LLM has historically been inconsistent — sometimes emitting `8.0` to
    mean $8M, other times `18072289.0` to mean $18M. Heuristic: values below
    1,000 are treated as millions (no real funding round is <$1K).
    """
    if value is None:
        return None
    try:
        v = float(value)
    except (TypeError, ValueError):
        return None
    if v <= 0:
        return None
    return v * 1_000_000 if v < 1000 else v


def _format_usd(usd: float | None) -> str:
    if usd is None:
        return ""
    if usd >= 1_000_000_000:
        return f"${usd / 1_000_000_000:.1f}B"
    if usd >= 1_000_000:
        return f"${usd / 1_000_000:.1f}M".replace(".0M", "M")
    if usd >= 1_000:
        return f"${usd / 1_000:.0f}K"
    return f"${usd:.0f}"


def _format_native_inr(amount: float) -> str:
    if amount >= 10_000_000:  # 1 crore = 10M
        crores = amount / 10_000_000
        return f"₹{crores:.0f} Cr" if crores >= 10 else f"₹{crores:.1f} Cr"
    if amount >= 100_000:  # 1 lakh = 100K
        return f"₹{amount / 100_000:.0f} L"
    return f"₹{amount:,.0f}"


def _format_amount(
    amount_usd: float | None,
    amount: float | None,
    currency: str | None,
) -> str:
    """Render a funding amount for humans.

    Rules:
    - USD rounds: "$15M" (derived from normalized amount_usd)
    - INR rounds: "₹150 Cr ($18M)" — native first, USD in parens
    - Other currencies: "5M EUR ($5.4M)"
    - Missing amount: ""
    """
    usd = _normalize_usd(amount_usd)
    usd_str = _format_usd(usd)

    if currency == "INR" and amount:
        return f"{_format_native_inr(amount)} ({usd_str})" if usd_str else _format_native_inr(amount)
    if currency and currency not in ("USD", None) and amount:
        native = f"{amount:,.0f} {currency}" if amount >= 1000 else f"{amount:.1f}M {currency}"
        return f"{native} ({usd_str})" if usd_str else native
    return usd_str


def _format_date(value: Any) -> str:
    """Normalize a date field to ISO YYYY-MM-DD. Drops time and microseconds."""
    if value in (None, ""):
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    s = str(value)
    # SQLite returns ISO strings; lop off time suffix if present
    return s.split("T", 1)[0].split(" ", 1)[0]


def _fetch_rows(storage: Storage, min_confidence: float = 0.0) -> list[dict[str, Any]]:
    """Raw DB join. Returns rows with DB-shape fields — not export-shape."""
    q = """
    SELECT
        fr.round_id,
        fr.company_name,
        fr.company_json,
        fr.stage,
        fr.amount,
        fr.currency,
        fr.amount_usd,
        fr.announced_on,
        fr.summary,
        fr.confidence,
        fr.extraction_method,
        fr.created_at,
        fr.updated_at,
        (SELECT GROUP_CONCAT(name, ' | ')
         FROM round_investors ri WHERE ri.round_id = fr.round_id) AS investors,
        (SELECT name FROM round_investors ri
         WHERE ri.round_id = fr.round_id AND ri.lead = 1 LIMIT 1) AS lead_investor,
        (SELECT GROUP_CONCAT(article_url, ' | ')
         FROM round_sources rs WHERE rs.round_id = fr.round_id) AS sources
    FROM funding_rounds fr
    WHERE fr.confidence >= ?
    ORDER BY fr.confidence DESC, fr.announced_on DESC
    """
    out: list[dict[str, Any]] = []
    with storage.connect() as c:
        for row in c.execute(q, (min_confidence,)).fetchall():
            d = dict(row)
            try:
                company = json.loads(d.pop("company_json") or "{}")
            except json.JSONDecodeError:
                company = {}
            d["sector"] = company.get("sector")
            d["city"] = company.get("city")
            out.append(d)
    return out


def build_public_rows(raw_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Shape DB rows into the clean public dict used across exporters + dashboard.

    Each row carries display-formatted strings (`stage`, `amount`) plus raw
    values needed by the dashboard's filters (`stage_raw`, `amount_usd_raw`).
    CSV/XLSX exporters read `PUBLIC_COLUMNS` only, so the raw extras are ignored
    unless the consumer asks for them.
    """
    out: list[dict[str, Any]] = []
    for r in raw_rows:
        stage_raw = (r.get("stage") or "").strip().lower() or None
        amount_usd_raw = _normalize_usd(r.get("amount_usd"))
        out.append({
            "company_name": r.get("company_name") or "",
            "sector": r.get("sector") or "",
            "stage": _format_stage(stage_raw),
            "stage_raw": stage_raw,
            "amount": _format_amount(r.get("amount_usd"), r.get("amount"), r.get("currency")),
            "amount_usd": round(amount_usd_raw, 2) if amount_usd_raw is not None else None,
            "amount_usd_raw": amount_usd_raw,
            "currency": r.get("currency") or "",
            "announced_on": _format_date(r.get("announced_on")),
            "lead_investor": r.get("lead_investor") or "",
            "investors": r.get("investors") or "",
            "sources": r.get("sources") or "",
            "confidence": round(float(r["confidence"]), 3) if r.get("confidence") is not None else 0.0,
            "summary": r.get("summary") or "",
            "city": r.get("city") or "",
        })
    return out


def _row_for_csv(r: dict[str, Any]) -> dict[str, Any]:
    """Pick only PUBLIC_COLUMNS and render None → ''."""
    return {k: ("" if r.get(k) is None else r.get(k)) for k in PUBLIC_COLUMNS}


def export_csv(storage: Storage, path: Path, min_confidence: float = 0.0) -> int:
    rows = build_public_rows(_fetch_rows(storage, min_confidence=min_confidence))
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=PUBLIC_COLUMNS)
        writer.writeheader()
        for r in rows:
            writer.writerow(_row_for_csv(r))
    return len(rows)


def export_json(storage: Storage, path: Path, min_confidence: float = 0.0) -> int:
    """JSON keeps arrays for investors/sources and adds raw fields for the dashboard."""
    rows = build_public_rows(_fetch_rows(storage, min_confidence=min_confidence))
    path.parent.mkdir(parents=True, exist_ok=True)
    out: list[dict[str, Any]] = []
    for r in rows:
        out.append({
            "company_name": r["company_name"],
            "sector": r["sector"],
            "stage": r["stage"],
            "stage_raw": r["stage_raw"],
            "amount": r["amount"],
            "amount_usd": r["amount_usd"],
            "currency": r["currency"],
            "announced_on": r["announced_on"],
            "lead_investor": r["lead_investor"],
            "investors": [x for x in r["investors"].split(" | ") if x] if r["investors"] else [],
            "sources": [x for x in r["sources"].split(" | ") if x] if r["sources"] else [],
            "confidence": r["confidence"],
            "summary": r["summary"],
            "city": r["city"],
        })
    with path.open("w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2, default=str)
    return len(out)


def export_xlsx(storage: Storage, path: Path, min_confidence: float = 0.0) -> int:
    """Excel workbook with one sheet 'rounds'. Frozen header, column widths,
    conditional formatting on confidence (red below 0.35, green at/above 0.85).
    """
    from openpyxl import Workbook
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    rows = build_public_rows(_fetch_rows(storage, min_confidence=min_confidence))
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "rounds"

    ws.append(PUBLIC_COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E8E8E8")
    ws.freeze_panes = "A2"

    for r in rows:
        ws.append([_xlsx_cell(r.get(c)) for c in PUBLIC_COLUMNS])

    for idx, col_name in enumerate(PUBLIC_COLUMNS, start=1):
        letter = get_column_letter(idx)
        max_len = max(
            [len(col_name)] + [len(str(r.get(col_name) or "")) for r in rows]
        )
        ws.column_dimensions[letter].width = min(max_len + 2, 50)

    conf_col = PUBLIC_COLUMNS.index("confidence") + 1
    conf_letter = get_column_letter(conf_col)
    if rows:
        rng = f"{conf_letter}2:{conf_letter}{len(rows) + 1}"
        ws.conditional_formatting.add(rng, CellIsRule(
            operator="lessThan", formula=["0.35"],
            fill=PatternFill("solid", fgColor="FADBD8"),
        ))
        ws.conditional_formatting.add(rng, CellIsRule(
            operator="greaterThanOrEqual", formula=["0.85"],
            fill=PatternFill("solid", fgColor="D5F5E3"),
        ))

    wb.save(path)
    return len(rows)


def _xlsx_cell(v: object) -> object:
    if v is None:
        return ""
    if isinstance(v, (int, float)):
        return v
    return str(v)
